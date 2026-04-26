from __future__ import annotations

import csv
import json
import os
import re
import subprocess
import zipfile
from dataclasses import asdict, dataclass
from io import BytesIO, StringIO
from typing import Optional, Tuple
from xml.etree import ElementTree

import fitz
from openpyxl import load_workbook

from app.document_files import file_extension


TEXT_JSON_FILENAME = "text.json"
DEFAULT_OCR_DPI_SCALE = 2.0
DEFAULT_OCR_MAX_PAGES = 25
MIN_EMBEDDED_TEXT_CHARS = 300
MIN_TEXT_QUALITY_SCORE = 0.55
MAX_OCR_SAMPLE_PAGES = 3


@dataclass
class ExtractedText:
    text: str
    status: str
    method: str
    error: Optional[str] = None
    warning: Optional[str] = None
    embedded_quality: Optional[float] = None
    ocr_quality: Optional[float] = None


def text_json_payload(
    *,
    document_id: str,
    original_filename: str,
    content_type: str,
    data: bytes,
    source: str,
) -> bytes:
    extracted = extract_document_text(original_filename, content_type, data)
    payload = {
        "document_id": document_id,
        "original_filename": original_filename,
        "content_type": content_type,
        "source": source,
        **asdict(extracted),
    }
    return json.dumps(payload, indent=2, sort_keys=True).encode("utf-8")


def extract_document_text(filename: str, content_type: str, data: bytes) -> ExtractedText:
    extension = file_extension(filename)
    try:
        if content_type == "application/pdf" or extension == ".pdf":
            return _extract_pdf_text(data)
        if content_type in {"text/plain", "text/csv"} or extension in {".txt", ".csv"}:
            return ExtractedText(text=_decode_text(data), status="extracted", method="direct")
        if (
            content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or extension == ".docx"
        ):
            return ExtractedText(text=_extract_docx_text(data), status="extracted", method="docx")
        if (
            content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or extension == ".xlsx"
        ):
            return ExtractedText(text=_extract_xlsx_text(data), status="extracted", method="xlsx")
        if content_type in {"image/png", "image/jpeg"} or extension in {".png", ".jpg", ".jpeg"}:
            text = _run_tesseract(data).strip()
            return ExtractedText(text=text, status="ocr_extracted", method="ocr")
    except Exception as error:
        return ExtractedText(text="", status="failed", method="none", error=str(error))

    return ExtractedText(text="", status="unsupported", method="none")


def _extract_pdf_text(data: bytes) -> ExtractedText:
    document = fitz.open(stream=data, filetype="pdf")
    embedded_text = _extract_pdf_embedded_text(document)
    embedded_quality = _text_quality_score(embedded_text)

    if _should_try_ocr(document, embedded_text, embedded_quality):
        try:
            ocr_text, warning = _ocr_pdf_text(document)
        except Exception as error:
            if embedded_text.strip():
                return ExtractedText(
                    text=embedded_text,
                    status="extracted",
                    method="embedded",
                    warning=f"OCR failed; using embedded PDF text: {error}",
                    embedded_quality=embedded_quality,
                )
            return ExtractedText(
                text="",
                status="failed",
                method="none",
                error=f"PDF text extraction produced no usable text and OCR failed: {error}",
                embedded_quality=embedded_quality,
            )

        ocr_quality = _text_quality_score(ocr_text)
        if _prefer_ocr(embedded_text, embedded_quality, ocr_text, ocr_quality):
            return ExtractedText(
                text=ocr_text.strip(),
                status="ocr_extracted",
                method="ocr",
                warning=warning,
                embedded_quality=embedded_quality,
                ocr_quality=ocr_quality,
            )

    if embedded_text.strip():
        return ExtractedText(
            text=embedded_text,
            status="extracted",
            method="embedded",
            embedded_quality=embedded_quality,
        )

    return ExtractedText(text="", status="failed", method="none", error="PDF extraction produced no text")


def _extract_pdf_embedded_text(document: fitz.Document) -> str:
    return "\n\n".join(page.get_text().strip() for page in document).strip()


def _should_try_ocr(document: fitz.Document, embedded_text: str, embedded_quality: float) -> bool:
    if not embedded_text.strip():
        return True
    if len(embedded_text) < MIN_EMBEDDED_TEXT_CHARS:
        return True
    return embedded_quality < MIN_TEXT_QUALITY_SCORE and _is_image_heavy_pdf(document)


def _is_image_heavy_pdf(document: fitz.Document) -> bool:
    sample_pages = min(len(document), MAX_OCR_SAMPLE_PAGES)
    if sample_pages == 0:
        return False

    image_heavy_pages = 0
    for page_index in range(sample_pages):
        page = document[page_index]
        page_area = max(page.rect.width * page.rect.height, 1)
        image_area = 0.0
        for image in page.get_images(full=True):
            for rect in page.get_image_rects(image[0]):
                image_area += rect.width * rect.height
        if image_area / page_area >= 0.5:
            image_heavy_pages += 1

    return image_heavy_pages >= max(1, sample_pages // 2)


def _ocr_pdf_text(document: fitz.Document) -> Tuple[str, Optional[str]]:
    max_pages = _ocr_max_pages()
    page_count = len(document)
    pages_to_process = min(page_count, max_pages) if max_pages else page_count
    matrix = fitz.Matrix(_ocr_dpi_scale(), _ocr_dpi_scale())
    page_text = []

    for page_index in range(pages_to_process):
        page = document[page_index]
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        page_text.append(_run_tesseract(pixmap.tobytes("png")).strip())

    warning = None
    if max_pages and page_count > max_pages:
        warning = f"OCR limited to first {max_pages} of {page_count} pages"

    return "\n\n".join(text for text in page_text if text), warning


def _prefer_ocr(
    embedded_text: str,
    embedded_quality: float,
    ocr_text: str,
    ocr_quality: float,
) -> bool:
    if not ocr_text.strip():
        return False
    if not embedded_text.strip():
        return True
    if len(embedded_text) < MIN_EMBEDDED_TEXT_CHARS:
        return True
    return ocr_quality >= embedded_quality + 0.08


def _text_quality_score(text: str) -> float:
    stripped = text.strip()
    if not stripped:
        return 0.0

    tokens = re.findall(r"[A-Za-z0-9][A-Za-z0-9'./@#%&+-]*", stripped)
    if not tokens:
        return 0.0

    printable_ratio = sum(character.isprintable() for character in stripped) / len(stripped)
    word_like = sum(_is_word_like(token) for token in tokens) / len(tokens)
    short_noise = sum(1 for token in tokens if len(token) == 1) / len(tokens)
    return max(0.0, min(1.0, (printable_ratio * 0.35) + (word_like * 0.75) - short_noise))


def _is_word_like(token: str) -> bool:
    if token.isdigit():
        return True
    letters = [character for character in token if character.isalpha()]
    if not letters:
        return True
    if len(letters) <= 2:
        return True
    if sum(character.lower() in "aeiouy" for character in letters) > 0:
        return True
    return token.isupper() and len(letters) <= 6


def _run_tesseract(image_data: bytes) -> str:
    command = os.getenv("DOCUMENT_OCR_TESSERACT_CMD", "tesseract")
    language = os.getenv("DOCUMENT_OCR_LANGUAGE", "eng")
    try:
        result = subprocess.run(
            [command, "stdin", "stdout", "-l", language],
            input=image_data,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
    except FileNotFoundError as error:
        raise RuntimeError(
            f"Tesseract command '{command}' is not installed or not on PATH"
        ) from error

    if result.returncode != 0:
        detail = result.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(detail or f"Tesseract exited with status {result.returncode}")

    return result.stdout.decode("utf-8", errors="replace")


def _ocr_max_pages() -> int:
    value = os.getenv("DOCUMENT_OCR_MAX_PAGES", str(DEFAULT_OCR_MAX_PAGES))
    try:
        return max(0, int(value))
    except ValueError:
        return DEFAULT_OCR_MAX_PAGES


def _ocr_dpi_scale() -> float:
    value = os.getenv("DOCUMENT_OCR_DPI_SCALE", str(DEFAULT_OCR_DPI_SCALE))
    try:
        return max(1.0, float(value))
    except ValueError:
        return DEFAULT_OCR_DPI_SCALE


def _extract_docx_text(data: bytes) -> str:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(BytesIO(data)) as archive:
        document_xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(document_xml)
    paragraphs = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        paragraph_text = "".join(texts).strip()
        if paragraph_text:
            paragraphs.append(paragraph_text)
    return "\n".join(paragraphs).strip()


def _extract_xlsx_text(data: bytes) -> str:
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    lines = []
    try:
        for worksheet in workbook.worksheets:
            lines.append(f"[{worksheet.title}]")
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    output = StringIO()
                    csv.writer(output).writerow(values)
                    lines.append(output.getvalue().strip())
    finally:
        workbook.close()
    return "\n".join(lines).strip()


def _decode_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace").strip()
