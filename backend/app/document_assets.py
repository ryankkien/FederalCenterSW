from __future__ import annotations

import csv
import json
import os
import re
import subprocess
import zipfile
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import simpleSplit
from reportlab.pdfgen import canvas

from app.blob_storage import BlobStorage
from app.document_files import file_extension

try:
    import fitz
except ImportError:
    fitz = None


TEXT_JSON_FILENAME = "text.json"
DEFAULT_OCR_DPI_SCALE = 2.0
DEFAULT_OCR_MAX_PAGES = 25
MIN_EMBEDDED_TEXT_CHARS = 300
MIN_TEXT_QUALITY_SCORE = 0.55
MAX_OCR_SAMPLE_PAGES = 3


@dataclass
class StoredContractDocument:
    blob_path: str
    text_blob_path: str
    stored_filename: str
    content_type: str
    size_bytes: int


def store_contract_document(
    storage: BlobStorage,
    document_id: str,
    original_filename: str,
    content_type: str,
    data: bytes,
    source: str,
) -> StoredContractDocument:
    converted = _convert_primary_file(original_filename, content_type, data)
    folder = f"contracts/{document_id}"
    blob_path = f"{folder}/{converted.filename}"
    text_blob_path = f"{folder}/{TEXT_JSON_FILENAME}"

    storage.upload_bytes(blob_path, converted.data, converted.content_type)
    storage.upload_bytes(
        text_blob_path,
        _text_json_payload(
            document_id=document_id,
            original_filename=original_filename,
            original_content_type=content_type,
            original_data=data,
            stored_filename=converted.filename,
            content_type=converted.content_type,
            source=source,
        ),
        "application/json",
    )

    return StoredContractDocument(
        blob_path=blob_path,
        text_blob_path=text_blob_path,
        stored_filename=converted.filename,
        content_type=converted.content_type,
        size_bytes=len(data),
    )


@dataclass
class _ConvertedFile:
    filename: str
    content_type: str
    data: bytes


def _convert_primary_file(filename: str, content_type: str, data: bytes) -> _ConvertedFile:
    extension = file_extension(filename)
    if extension == ".pdf":
        return _ConvertedFile(filename="main.pdf", content_type="application/pdf", data=data)
    if extension in {".txt", ".csv"}:
        try:
            text = _decode_text(data)
            return _ConvertedFile(
                filename="main.pdf",
                content_type="application/pdf",
                data=_text_to_pdf(text),
            )
        except Exception:
            return _ConvertedFile(filename=f"main{extension}", content_type=content_type, data=data)
    if extension in {".png", ".jpg", ".jpeg"}:
        try:
            return _ConvertedFile(
                filename="main.pdf",
                content_type="application/pdf",
                data=_image_to_pdf(data),
            )
        except Exception:
            return _ConvertedFile(filename=f"main{extension}", content_type=content_type, data=data)
    return _ConvertedFile(filename=f"main{extension or '.bin'}", content_type=content_type, data=data)


def _text_json_payload(
    document_id: str,
    original_filename: str,
    original_content_type: str,
    original_data: bytes,
    stored_filename: str,
    content_type: str,
    source: str,
) -> bytes:
    extracted = _extract_text(original_filename, original_content_type, original_data)
    payload = {
        "document_id": document_id,
        "original_filename": original_filename,
        "stored_filename": stored_filename,
        "content_type": content_type,
        "source": source,
        "text": extracted.text,
        "status": extracted.status,
        "method": extracted.method,
        "extraction_status": extracted.status,
        "extraction_error": extracted.error,
        "extraction_warning": extracted.warning,
        "embedded_quality": extracted.embedded_quality,
        "ocr_quality": extracted.ocr_quality,
        "pages": extracted.pages,
    }
    return json.dumps(payload, indent=2, sort_keys=True).encode("utf-8")


@dataclass
class _ExtractedText:
    text: str
    status: str
    method: str = "none"
    error: Optional[str] = None
    warning: Optional[str] = None
    embedded_quality: Optional[float] = None
    ocr_quality: Optional[float] = None
    pages: Optional[List[Dict[str, object]]] = None

    def __post_init__(self) -> None:
        if self.pages is None:
            self.pages = []


def _extract_text(filename: str, content_type: str, data: bytes) -> _ExtractedText:
    extension = file_extension(filename)
    try:
        if content_type == "application/pdf" or extension == ".pdf":
            return _extract_pdf_text(data)
        elif content_type in {"text/plain", "text/csv"} or extension in {".txt", ".csv"}:
            text = _decode_text(data)
            method = "direct"
        elif content_type in {"image/png", "image/jpeg"} or extension in {".png", ".jpg", ".jpeg"}:
            text, warning = _ocr_image_text(data)
            return _ExtractedText(
                text=text,
                status="ocr_extracted",
                method="ocr",
                warning=warning,
                pages=_single_page_payload(text, "ocr_extracted", warning=warning),
            )
        elif (
            content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or extension == ".docx"
        ):
            text = _extract_docx_text(data)
            method = "docx"
        elif (
            content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or extension == ".xlsx"
        ):
            text = _extract_xlsx_text(data)
            method = "xlsx"
        else:
            return _ExtractedText(text="", status="unsupported")
    except Exception as error:
        return _ExtractedText(text="", status="failed", error=str(error))

    return _ExtractedText(
        text=text,
        status="extracted",
        method=method,
        pages=_single_page_payload(text, "extracted"),
    )


def _extract_pdf_text(data: bytes) -> _ExtractedText:
    embedded_text, pages, document = _extract_pdf_embedded_text(data)
    embedded_quality = _text_quality_score(embedded_text)

    if _should_try_ocr(document, embedded_text, embedded_quality):
        try:
            ocr_text, warning = _ocr_pdf_text(data)
        except Exception as error:
            if embedded_text.strip():
                return _ExtractedText(
                    text=embedded_text,
                    status="extracted",
                    method="embedded",
                    warning=f"OCR failed; using embedded PDF text: {error}",
                    embedded_quality=embedded_quality,
                    pages=pages,
                )
            return _ExtractedText(
                text="",
                status="failed",
                method="none",
                error=f"PDF text extraction produced no usable text and OCR failed: {error}",
                embedded_quality=embedded_quality,
            )

        ocr_quality = _text_quality_score(ocr_text)
        if _prefer_ocr(embedded_text, embedded_quality, ocr_text, ocr_quality):
            return _ExtractedText(
                text=ocr_text.strip(),
                status="ocr_extracted",
                method="ocr",
                warning=warning,
                embedded_quality=embedded_quality,
                ocr_quality=ocr_quality,
                pages=_single_page_payload(ocr_text.strip(), "ocr_extracted", warning=warning),
            )

    if embedded_text.strip():
        return _ExtractedText(
            text=embedded_text,
            status="extracted",
            method="embedded",
            embedded_quality=embedded_quality,
            pages=pages,
        )

    return _ExtractedText(
        text="",
        status="failed",
        method="none",
        error="PDF extraction produced no text",
        embedded_quality=embedded_quality,
    )


def _extract_pdf_embedded_text(
    data: bytes,
) -> Tuple[str, List[Dict[str, object]], Optional[object]]:
    if fitz is not None:
        document = fitz.open(stream=data, filetype="pdf")
        pages = []
        page_text_values = []
        offset = 0
        for index, page in enumerate(document, start=1):
            page_text = (page.get_text() or "").strip()
            if not page_text:
                continue
            start = offset
            page_text_values.append(page_text)
            offset += len(page_text)
            pages.append(
                {
                    "page_number": index,
                    "text": page_text,
                    "start_char": start,
                    "end_char": offset,
                    "extraction_status": "extracted",
                }
            )
            offset += 2
        return "\n\n".join(page_text_values).strip(), pages, document

    reader = PdfReader(BytesIO(data))
    pages = []
    offset = 0
    page_text_values = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        if page_text:
            start = offset
            page_text_values.append(page_text)
            offset += len(page_text)
            pages.append(
                {
                    "page_number": index,
                    "text": page_text,
                    "start_char": start,
                    "end_char": offset,
                    "extraction_status": "extracted",
                }
            )
            offset += 2
    return "\n\n".join(page_text_values).strip(), pages, None


def _should_try_ocr(
    document: Optional[object],
    embedded_text: str,
    embedded_quality: float,
) -> bool:
    if not embedded_text.strip():
        return True
    if len(embedded_text) < MIN_EMBEDDED_TEXT_CHARS:
        return True
    return bool(document) and embedded_quality < MIN_TEXT_QUALITY_SCORE and _is_image_heavy_pdf(document)


def _is_image_heavy_pdf(document: object) -> bool:
    sample_pages = min(len(document), MAX_OCR_SAMPLE_PAGES)
    if sample_pages == 0:
        return False

    image_heavy_pages = 0
    for page_index in range(sample_pages):
        page = document[page_index]
        page_area = max(page.rect.width * page.rect.height, 1)
        image_area = 0.0
        for image in page.get_images(full=True):
            for rect in page.get_image_rects(image[0]):
                image_area += rect.width * rect.height
        if image_area / page_area >= 0.5:
            image_heavy_pages += 1

    return image_heavy_pages >= max(1, sample_pages // 2)


def _prefer_ocr(
    embedded_text: str,
    embedded_quality: float,
    ocr_text: str,
    ocr_quality: float,
) -> bool:
    if not ocr_text.strip():
        return False
    if not embedded_text.strip():
        return True
    if len(embedded_text) < MIN_EMBEDDED_TEXT_CHARS:
        return True
    return ocr_quality >= embedded_quality + 0.08


def _text_quality_score(text: str) -> float:
    stripped = text.strip()
    if not stripped:
        return 0.0

    tokens = re.findall(r"[A-Za-z0-9][A-Za-z0-9'./@#%&+-]*", stripped)
    if not tokens:
        return 0.0

    printable_ratio = sum(character.isprintable() for character in stripped) / len(stripped)
    word_like = sum(_is_word_like(token) for token in tokens) / len(tokens)
    short_noise = sum(1 for token in tokens if len(token) == 1) / len(tokens)
    return max(0.0, min(1.0, (printable_ratio * 0.35) + (word_like * 0.75) - short_noise))


def _is_word_like(token: str) -> bool:
    if token.isdigit():
        return True
    letters = [character for character in token if character.isalpha()]
    if not letters:
        return True
    if len(letters) <= 2:
        return True
    if sum(character.lower() in "aeiouy" for character in letters) > 0:
        return True
    return token.isupper() and len(letters) <= 6


def _ocr_pdf_text(data: bytes) -> Tuple[str, Optional[str]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF is required for PDF OCR rendering")

    document = fitz.open(stream=data, filetype="pdf")
    max_pages = _ocr_max_pages()
    page_count = len(document)
    pages_to_process = min(page_count, max_pages) if max_pages else page_count
    scale = _ocr_dpi_scale()
    matrix = fitz.Matrix(scale, scale)
    page_text = []

    for page_index in range(pages_to_process):
        page = document[page_index]
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        page_text.append(_run_tesseract(pixmap.tobytes("png")).strip())

    warning = None
    if max_pages and page_count > max_pages:
        warning = f"OCR limited to first {max_pages} of {page_count} pages"

    return "\n\n".join(text for text in page_text if text), warning


def _ocr_image_text(data: bytes) -> Tuple[str, Optional[str]]:
    text = _run_tesseract(data).strip()
    if not text:
        raise RuntimeError("image OCR completed but produced no text")
    return text, None


def _single_page_payload(
    text: str,
    status: str,
    warning: Optional[str] = None,
) -> List[Dict[str, object]]:
    if not text:
        return []
    payload: Dict[str, object] = {
        "page_number": 1,
        "text": text,
        "start_char": 0,
        "end_char": len(text),
        "extraction_status": status,
    }
    if warning:
        payload["extraction_warning"] = warning
    return [payload]


def _run_tesseract(image_data: bytes) -> str:
    command = os.getenv("DOCUMENT_OCR_TESSERACT_CMD", "tesseract")
    language = os.getenv("DOCUMENT_OCR_LANGUAGE", "eng")
    try:
        result = subprocess.run(
            [command, "stdin", "stdout", "-l", language],
            input=image_data,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
    except FileNotFoundError as error:
        raise RuntimeError(
            f"Tesseract command '{command}' is not installed or not on PATH"
        ) from error

    if result.returncode != 0:
        detail = result.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(detail or f"Tesseract exited with status {result.returncode}")

    return result.stdout.decode("utf-8", errors="replace")


def _ocr_max_pages() -> int:
    value = os.getenv("DOCUMENT_OCR_MAX_PAGES", str(DEFAULT_OCR_MAX_PAGES))
    try:
        return max(0, int(value))
    except ValueError:
        return DEFAULT_OCR_MAX_PAGES


def _ocr_dpi_scale() -> float:
    value = os.getenv("DOCUMENT_OCR_DPI_SCALE", str(DEFAULT_OCR_DPI_SCALE))
    try:
        return max(1.0, float(value))
    except ValueError:
        return DEFAULT_OCR_DPI_SCALE


def _extract_docx_text(data: bytes) -> str:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(BytesIO(data)) as archive:
        document_xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(document_xml)
    paragraphs = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        paragraph_text = "".join(texts).strip()
        if paragraph_text:
            paragraphs.append(paragraph_text)
    return "\n".join(paragraphs).strip()


def _extract_xlsx_text(data: bytes) -> str:
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    lines = []
    for worksheet in workbook.worksheets:
        lines.append(f"[{worksheet.title}]")
        for row in worksheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                output = StringIO()
                csv.writer(output).writerow(values)
                lines.append(output.getvalue().strip())
    workbook.close()
    return "\n".join(lines).strip()


def _decode_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace").strip()


def _text_to_pdf(text: str) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 72
    y = height - margin
    line_height = 14

    for paragraph in text.splitlines() or [""]:
        lines = simpleSplit(paragraph or " ", "Helvetica", 10, width - (margin * 2))
        for line in lines:
            if y < margin:
                pdf.showPage()
                y = height - margin
            pdf.drawString(margin, y, line)
            y -= line_height
        if y < margin:
            pdf.showPage()
            y = height - margin
        y -= line_height

    pdf.save()
    return buffer.getvalue()


def _image_to_pdf(data: bytes) -> bytes:
    with Image.open(BytesIO(data)) as image:
        converted = image.convert("RGB")
        buffer = BytesIO()
        converted.save(buffer, format="PDF")
        return buffer.getvalue()
